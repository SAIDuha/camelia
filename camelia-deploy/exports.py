"""
Camélia — Export Module
Generates Excel (.xlsx) and PDF reports for employee work hours.
"""
import io
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

WEEKDAYS_FR = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
MONTHS_FR = ["Janvier","Février","Mars","Avril","Mai","Juin","Juillet","Août","Septembre","Octobre","Novembre","Décembre"]

def calc_work_minutes(badge):
    if not badge['arrival_time'] or not badge['departure_time']:
        return None
    ah, am = map(int, badge['arrival_time'].split(':'))
    dh, dm = map(int, badge['departure_time'].split(':'))
    total = (dh * 60 + dm) - (ah * 60 + am)
    if badge.get('break_start') and badge.get('break_end'):
        bsh, bsm = map(int, badge['break_start'].split(':'))
        beh, bem = map(int, badge['break_end'].split(':'))
        total -= (beh * 60 + bem) - (bsh * 60 + bsm)
    return max(0, total)

def fmt_hours(mins):
    if mins is None:
        return "—"
    h = mins // 60
    m = mins % 60
    return f"{h}h{m:02d}"

def fmt_date_fr(d):
    if isinstance(d, str):
        parts = d.split('-')
        d = date(int(parts[0]), int(parts[1]), int(parts[2]))
    return f"{WEEKDAYS_FR[d.weekday()]} {d.day} {MONTHS_FR[d.month-1]} {d.year}"

def get_period_label(period, start_date, end_date):
    if period == 'week':
        return f"Semaine du {start_date.day}/{start_date.month} au {end_date.day}/{end_date.month}/{end_date.year}"
    elif period == 'month':
        return f"{MONTHS_FR[start_date.month-1]} {start_date.year}"
    else:
        return f"Du {start_date.day}/{start_date.month}/{start_date.year} au {end_date.day}/{end_date.month}/{end_date.year}"

# ═══════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════
def generate_excel(company_name, employees_data, period_label):
    """
    Generates an Excel workbook with one sheet per employee + a summary sheet.
    employees_data = [{ 'name': str, 'code': str, 'department': str, 'badges': [badge_dicts] }, ...]
    Returns bytes.
    """
    wb = Workbook()

    # Colors
    rose = "B5577A"
    rose_light = "F8EEF1"
    dark = "2C2825"
    gray = "9A938C"
    green = "3AAB6D"
    border_color = "E8E3DD"

    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=rose, end_color=rose, fill_type="solid")
    title_font = Font(name='Calibri', size=14, bold=True, color=dark)
    subtitle_font = Font(name='Calibri', size=10, color=gray)
    data_font = Font(name='Calibri', size=10, color=dark)
    total_font = Font(name='Calibri', size=11, bold=True, color=rose)
    thin_border = Border(
        bottom=Side(style='thin', color=border_color)
    )

    # ── Summary Sheet ──
    ws_summary = wb.active
    ws_summary.title = "Résumé"
    ws_summary.sheet_properties.tabColor = rose

    ws_summary.merge_cells('A1:F1')
    ws_summary['A1'] = f"🌸 Camélia — {company_name}"
    ws_summary['A1'].font = Font(name='Calibri', size=16, bold=True, color=rose)
    ws_summary['A1'].alignment = Alignment(horizontal='center')

    ws_summary.merge_cells('A2:F2')
    ws_summary['A2'] = period_label
    ws_summary['A2'].font = subtitle_font
    ws_summary['A2'].alignment = Alignment(horizontal='center')

    summary_headers = ['Employé', 'Code', 'Département', 'Jours travaillés', 'Heures totales', 'Moyenne/jour']
    for col, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    row = 5
    for emp in employees_data:
        badges = emp['badges']
        total_mins = 0
        days = 0
        for b in badges:
            w = calc_work_minutes(b)
            if w is not None:
                total_mins += w
                days += 1
        avg = total_mins // days if days else 0

        ws_summary.cell(row=row, column=1, value=emp['name']).font = data_font
        ws_summary.cell(row=row, column=2, value=emp['code']).font = data_font
        ws_summary.cell(row=row, column=3, value=emp['department']).font = data_font
        ws_summary.cell(row=row, column=4, value=days).font = data_font
        ws_summary.cell(row=row, column=5, value=fmt_hours(total_mins)).font = total_font
        ws_summary.cell(row=row, column=6, value=fmt_hours(avg)).font = data_font

        for col in range(1, 7):
            ws_summary.cell(row=row, column=col).border = thin_border
            ws_summary.cell(row=row, column=col).alignment = Alignment(horizontal='center')

        row += 1

    # Auto-width
    for col in range(1, 7):
        ws_summary.column_dimensions[get_column_letter(col)].width = 18

    # ── Per-employee sheets ──
    for emp in employees_data:
        safe_name = emp['name'][:28].replace('/', '-')
        ws = wb.create_sheet(title=safe_name)
        ws.sheet_properties.tabColor = rose

        ws.merge_cells('A1:G1')
        ws['A1'] = f"{emp['name']} — {emp['code']}"
        ws['A1'].font = title_font

        ws.merge_cells('A2:G2')
        ws['A2'] = f"{emp['department']} · {period_label}"
        ws['A2'].font = subtitle_font

        detail_headers = ['Date', 'Jour', 'Arrivée', 'Début pause', 'Fin pause', 'Départ', 'Total']
        for col, h in enumerate(detail_headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        row = 5
        grand_total = 0
        for b in sorted(emp['badges'], key=lambda x: x.get('date') or x.get('badge_date', '')):
            badge_date = b.get('date') or b.get('badge_date', '')
            if isinstance(badge_date, str) and badge_date:
                parts = badge_date.split('-')
                d = date(int(parts[0]), int(parts[1]), int(parts[2]))
            else:
                continue

            w = calc_work_minutes(b)
            if w is not None:
                grand_total += w

            ws.cell(row=row, column=1, value=f"{d.day}/{d.month}/{d.year}").font = data_font
            ws.cell(row=row, column=2, value=WEEKDAYS_FR[d.weekday()][:3]).font = data_font
            ws.cell(row=row, column=3, value=b.get('arrival_time') or '—').font = data_font
            ws.cell(row=row, column=4, value=b.get('break_start') or '—').font = data_font
            ws.cell(row=row, column=5, value=b.get('break_end') or '—').font = data_font
            ws.cell(row=row, column=6, value=b.get('departure_time') or '—').font = data_font

            total_cell = ws.cell(row=row, column=7, value=fmt_hours(w))
            if w is not None and w >= 420:
                total_cell.font = Font(name='Calibri', size=10, bold=True, color=green)
            elif w is not None:
                total_cell.font = Font(name='Calibri', size=10, color="D4973A")
            else:
                total_cell.font = Font(name='Calibri', size=10, color=gray)

            for col in range(1, 8):
                ws.cell(row=row, column=col).border = thin_border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')

            row += 1

        # Total row
        row += 1
        ws.cell(row=row, column=5, value="TOTAL").font = total_font
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=7, value=fmt_hours(grand_total)).font = total_font
        ws.cell(row=row, column=7).alignment = Alignment(horizontal='center')

        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ═══════════════════════════════════════════
# PDF EXPORT
# ═══════════════════════════════════════════
def generate_pdf(company_name, employees_data, period_label):
    """
    Generates a PDF report with all employees' hours.
    Returns bytes.
    """
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=20*mm, bottomMargin=15*mm, leftMargin=15*mm, rightMargin=15*mm)

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='CamTitle', fontName='Helvetica-Bold', fontSize=18, textColor=colors.HexColor('#B5577A'), alignment=TA_CENTER, spaceAfter=4*mm))
    styles.add(ParagraphStyle(name='CamSubtitle', fontName='Helvetica', fontSize=10, textColor=colors.HexColor('#9A938C'), alignment=TA_CENTER, spaceAfter=8*mm))
    styles.add(ParagraphStyle(name='CamEmpTitle', fontName='Helvetica-Bold', fontSize=12, textColor=colors.HexColor('#2C2825'), spaceBefore=6*mm, spaceAfter=3*mm))
    styles.add(ParagraphStyle(name='CamEmpSub', fontName='Helvetica', fontSize=9, textColor=colors.HexColor('#9A938C'), spaceAfter=3*mm))
    styles.add(ParagraphStyle(name='CamTotal', fontName='Helvetica-Bold', fontSize=10, textColor=colors.HexColor('#B5577A'), alignment=TA_RIGHT, spaceBefore=2*mm))

    elements = []

    # Header
    elements.append(Paragraph(f"🌸 Camélia — {company_name}", styles['CamTitle']))
    elements.append(Paragraph(f"Rapport de pointage · {period_label}", styles['CamSubtitle']))

    rose = colors.HexColor('#B5577A')
    rose_light = colors.HexColor('#F8EEF1')
    border_color = colors.HexColor('#E8E3DD')
    text_color = colors.HexColor('#2C2825')
    muted_color = colors.HexColor('#9A938C')
    green_color = colors.HexColor('#3AAB6D')
    warning_color = colors.HexColor('#D4973A')

    header_style = [
        ('BACKGROUND', (0, 0), (-1, 0), rose),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, 0), 0.5, rose),
        ('LINEBELOW', (0, 1), (-1, -2), 0.3, border_color),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, rose_light]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]

    for emp in employees_data:
        elements.append(Paragraph(f"{emp['name']} — {emp['code']}", styles['CamEmpTitle']))
        elements.append(Paragraph(f"{emp['department']}", styles['CamEmpSub']))

        table_data = [['Date', 'Jour', 'Arrivée', 'Pause début', 'Pause fin', 'Départ', 'Total']]

        grand_total = 0
        for b in sorted(emp['badges'], key=lambda x: x.get('date') or x.get('badge_date', '')):
            badge_date = b.get('date') or b.get('badge_date', '')
            if isinstance(badge_date, str) and badge_date:
                parts = badge_date.split('-')
                d = date(int(parts[0]), int(parts[1]), int(parts[2]))
            else:
                continue

            w = calc_work_minutes(b)
            if w is not None:
                grand_total += w

            table_data.append([
                f"{d.day}/{d.month}",
                WEEKDAYS_FR[d.weekday()][:3],
                b.get('arrival_time') or '—',
                b.get('break_start') or '—',
                b.get('break_end') or '—',
                b.get('departure_time') or '—',
                fmt_hours(w),
            ])

        if len(table_data) > 1:
            col_widths = [55, 40, 50, 55, 55, 50, 50]
            t = Table(table_data, colWidths=col_widths)
            style = TableStyle(header_style)

            # Color total column based on hours
            for i, b in enumerate(sorted(emp['badges'], key=lambda x: x.get('date') or x.get('badge_date', '')), 1):
                w = calc_work_minutes(b)
                if w is not None and w >= 420:
                    style.add('TEXTCOLOR', (6, i), (6, i), green_color)
                    style.add('FONTNAME', (6, i), (6, i), 'Helvetica-Bold')
                elif w is not None:
                    style.add('TEXTCOLOR', (6, i), (6, i), warning_color)

            t.setStyle(style)
            elements.append(t)

        elements.append(Paragraph(f"Total : {fmt_hours(grand_total)}", styles['CamTotal']))
        elements.append(Spacer(1, 4*mm))

    # Footer
    elements.append(Spacer(1, 8*mm))
    elements.append(Paragraph(
        f"Généré par Camélia · {date.today().strftime('%d/%m/%Y')}",
        ParagraphStyle(name='Footer', fontName='Helvetica', fontSize=8, textColor=muted_color, alignment=TA_CENTER)
    ))

    doc.build(elements)
    output.seek(0)
    return output.getvalue()
